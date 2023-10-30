import pandas as pd
import requests
import argparse
from datetime import datetime, timedelta
import yaml
from openpyxl import Workbook
from openpyxl.styles import PatternFill,Font

with open("config.yml","r") as file:
    init_settings = yaml.safe_load(file)

def arrange_excel_columns(keys):
    """
    Arrange columns in Excel according to keys args
    :param keys:
    :return:
    """
    excel_columns_template = ["rnr", "gruppe", "hu"]
    if keys is not None:
        excel_columns_template += keys

    return excel_columns_template


def find_colored(keys):
    if keys == ("false" or "False"):
        return False
    return True


def my_excel(columns, response_data, colored):
    """
    Writing given data into Excel file and coloring if needed
    :param columns:
    :param response_data:
    :param colored:
    """
    wb = Workbook()
    ws = wb.active

    date_now = datetime.now().date()

    # Columns Headers Fonts
    font_header = Font(bold=True)
    for col, col_data in enumerate(columns):
        cell = ws.cell(row=1, column=col + 1, value=col_data)
        cell.font = font_header

    # All data writing into Excel
    for row_index, row_data in enumerate(response_data.iterrows(), start=1):
        if colored and row_data[1]["hu"] != "":
            # differentiating dates in days
            date_str = row_data[1]["hu"]
            date_format = '%Y-%m-%d'
            date_obj = datetime.strptime(date_str, date_format).date()
            date_diff = date_now - date_obj

            colors = init_settings["client"]["colors"]
            if date_diff <= timedelta(days=30):
                row_color = colors["green"]
            elif date_diff <= timedelta(days=365):
                row_color = colors["orange"]
            else:
                row_color = colors["red"]

            # coloring rows with row color
            for cell in ws[row_index + 1]:
                cell.fill = PatternFill(
                    fill_type="solid",
                    start_color=row_color,
                    end_color=row_color
                )

        for col_index, col_data in enumerate(columns):
            cell_data = row_data[1][col_data]
            ws.cell(row=row_index + 1, column=col_index + 1, value=cell_data)


    wb.save(f"vehicles_{date_now}.xlsx")


if __name__ == "__main__":
    file_path = init_settings["client"]["file_path"]
    with open(file_path, "rb") as file:
        files = {"file": (file_path, file, "multipart/form-data")}
        response = requests.post(init_settings["client"]["server_url"], files=files)

    if response.status_code == 200:
        # argparsing -k (multiple input) and -c
        parser = argparse.ArgumentParser()
        parser.add_argument("-k", "--keys", help="list of keys", nargs="+")
        parser.add_argument("-c", "--colored", default=True, help="Flag to enable coloring")
        args = parser.parse_args()

        # Excel columns names and checking -c if it is colored
        excel_columns = arrange_excel_columns(args.keys)
        is_colored = find_colored(args.colored)

        # taking input and sorting by gruppe
        api_response = response.json()
        data = pd.DataFrame(api_response)
        data = data.sort_values(by="gruppe")

        my_excel(excel_columns, data, is_colored)
    else:
        print(f"request is failed status code = {response.status_code}: {response.text}")
